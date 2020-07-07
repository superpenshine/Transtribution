'''
Provide services for api views, 
db ops should be treated as models querysets
'''
from home.models import Grade
from django.conf import settings
from Serializers.GradeSerializer import GradeListSerializer

# File parse
import openpyxl
import numpy as np

# File preparation
from tempfile import NamedTemporaryFile

# Email
import os
import asyncio
import smtplib
from concurrent.futures import ThreadPoolExecutor
from email import encoders
from email.headerregistry import Address
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import COMMASPACE

# Grade Analytics
from django.db.models import Count, Avg, Max, Min, Q, F

_DEFAULT_POOL = ThreadPoolExecutor()

'''
Python3.2+, return future obj, use 'future obj'.result() to get return values
'''
def threaded(thread_pool=_DEFAULT_POOL):
    def decorator(func, *args, **kwargs):
        def inner(*args, **kwargs):
            return thread_pool.submit(func, *args, **kwargs)
        return inner
    return decorator

'''
Send email to address with given subj and content
smtp server and credentials must be set as env variables in advance.
to: aaa@aaa.aaa or [aaa@aaa.aaa, ...]

NOTE: django.core.mail is an easier choice here.
'''
@threaded()
def sendEmail(to, subject='Email from Transtribution', text='', files=[]):
    if isinstance(to, str): to = [to]
    if isinstance(files, str): files = [files]
    msg = MIMEMultipart()
    msg['From'] = settings.SMTP_HOST_USER
    msg['To'] = COMMASPACE.join(to)
    msg['Subject'] = subject
    msg.attach(MIMEText(text))

    for file in files:
        with open(file, 'rb') as f:
            part = MIMEApplication(f.read(), Name=os.path.basename(file))
        part['Content-Disposition'] = f"attachment; filename={os.path.basename(file)}"
        msg.attach(part)
    try:
        with smtplib.SMTP_SSL(settings.SMTP_HOST_ADDR, port=settings.SMTP_HOST_PORT) as smp:
            smp.login(settings.SMTP_HOST_USER, settings.SMTP_HOST_PWD)
            smp.send_message(msg)

    except Exception as e:
        for file in files: os.remove(file)
        return e

    for file in files: os.remove(file)

'''
Return a temporary xlsx file with given data
Input: [{a: 1, b:3}, {a:2, b:4}, ...]

'''
def createTmpFile(data, prefix='', suffix=''):
    wb = openpyxl.Workbook()
    tmp = NamedTemporaryFile(prefix=prefix, suffix=suffix, delete=False)

    ws = wb.active
    col_names = list(data[0].keys()) if data else []
    ws.append(col_names)
    for row in data:
        ws.append(list(row.values()))

    wb.save(tmp.name)
    # tmp.seek(0)
    return tmp.name

# Student identity attributes
pri_key_en = {
    '班级': 'class_name', 
    '学号': 'student_id', 
        '学生号': 'student_id', 
    '姓名': 'name', 
        '名字': 'name', 
        'name': 'name', 
    '密码': 'password', # for new student creation
        '身份证': 'password', 
        'pwd': 'password',
        'password': 'password',
}
# Grades attributes
key_en = {
    '成绩': 'score', 
        'grade': 'score', 
        'score': 'score', 
    '学科': 'subject', 
        '科目': 'subject', 
        '课程': 'subject', 
    '考试': 'test', 
}

'''
Find row # which contains colum info, col # which corresponds to grade/student attributes

key_order = {'name':{'class_name': xxx, 'student_id': xxx, ...}, 
            'score': xx, 
            'subject': xx
            'test': xx}
'''
def get_key_order(s):
    key_order = {'name': {}}
    header_row = 1
    for i, r in enumerate(s.iter_rows(values_only=True), 1):
        if len(key_order['name']) > 0:
            break
        header_row = i
        # Converting to np array later, start at 0
        for i, c in enumerate(r, 0):
            if c in key_en.keys():
                key_order[key_en[c]] = i 
            elif c in pri_key_en.keys():
                key_order['name'][pri_key_en[c]] = i

    if not key_order['name']:
        raise ValueError("Cannot find header line.")

    return key_order, header_row

'''
Get student/grade info from file
s_data: student attributes
g_data: grade attributes
'''
def parse(file):
    f = openpyxl.load_workbook(file)
    sheetNames = f.get_sheet_names()
    data = []

    for s_name in sheetNames:
        s = f[s_name]
        # Separate student and grade model data
        key_order, header_row = get_key_order(s)
        s_key_order = key_order.pop('name')
        g_key_order = key_order
        s_data = np.array([list(row) for row in s.iter_rows(min_row=header_row+1, values_only=True) if any(row)])[:, list(s_key_order.values())]
        g_data = np.array([list(row) for row in s.iter_rows(min_row=header_row+1, values_only=True) if any(row)])[:, list(g_key_order.values())]
        
        # Pack data
        for g_r, s_r in zip(g_data, s_data):
            _dict = {'name':{}}
            for g_k, g_v in zip(g_key_order.keys(), g_r):
                _dict[g_k] = g_v
            for s_k, s_v in zip(s_key_order.keys(), s_r):
                _dict['name'][s_k] = s_v
            data.append(_dict)

    return data

def handelFileSubmit(file):
    try:
        data = parse(file)
    except ValueError as e:
                   
        return str(e)

    # Serialize and validate
    ser = GradeListSerializer(data=data)
    if ser.is_valid():
        ser.save()

    return ser.errors


'''
Return all data if user is staff, 
else return user grades and class grades stats
'''
def getUserGrades(user, **extra_fields):
    qset = Grade.objects.filter(**extra_fields)
    if not user.is_staff:
        qset = qset.filter(name=user)
    grade_data = list(qset.values(
                            'test', 'subject', 'score', 'id', 
                            student_name=F('name_id__name'), 
                            password=F('name_id__password'), 
                            student_id=F('name_id__student_id'), 
                            class_name=F('name_id__class_name')
                            ))

    if not user.is_staff:
        class_grades = Grade.objects.filter(name__class_name=user.class_name)
        for grade in grade_data:
            subject, test, score = grade['subject'], grade['test'], grade['score']
            # Get ranking #
            grade.update(class_grades.aggregate(rank=Count('id', filter=Q(subject=subject, test=test, score__gte=score))))
            # Get test stats
            grade.update(class_grades.get_test_stats(subject, test, pass_grade=60))
            grade['avg'] = round(grade['avg'], 1)

    return grade_data